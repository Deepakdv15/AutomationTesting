import openpyxl

# wk=openpyxl.load_workbook(r"C:\python\RobotFramework1\DataSheet.xlsx")
# print(wk.sheetnames)
# activeSheetName=wk.active.title
# print(activeSheetName)
# active=wk["Sheet1"]
# print(active)
# print(active["B2"].value)
# print(active["A3"].value)
# cell1=active.cell(3,2)
# print(cell1.value)
# print(cell1.row)
# print(cell1.column)

workBook=openpyxl.Workbook()
print(workBook.sheetnames)
print(workBook.active.title)
sh=workBook.active
workBook.active.title="DeepakSingh"
sh1=workBook["DeepakSingh"]
sh1["A3"].value="testingndia"
workBook.create_sheet("ravi")
sh2=workBook["ravi"]
sh2["B4"].value=65447358976
workBook.remove(workBook["DeepakSingh"])
workBook.save(r"C:\python\RobotFramework1\datasheet1.xlsx")
